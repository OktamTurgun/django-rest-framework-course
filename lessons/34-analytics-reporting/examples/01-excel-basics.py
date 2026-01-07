"""
Excel Export Basics - openpyxl
Bu misolda:
- Excel workbook yaratish
- Ma'lumotlarni yozish
- Formatting va styling
- Formulas qo'shish
- Auto-sizing columns
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

print("=" * 70)
print("EXCEL EXPORT BASICS")
print("=" * 70)

# ============================================================================
# EXAMPLE 1: Basic Excel Creation
# ============================================================================
print("\nEXAMPLE 1: Creating Basic Excel File")
print("-" * 70)

# Create workbook
wb = Workbook()
ws = wb.active
ws.title = "Books Report"

# Sample data
books = [
    ["Title", "Author", "Price", "Stock", "Category"],
    ["Python Programming", "John Doe", 45.99, 10, "Programming"],
    ["Django for Beginners", "Jane Smith", 39.99, 5, "Web Development"],
    ["Clean Code", "Robert Martin", 49.99, 8, "Programming"],
    ["Design Patterns", "Gang of Four", 54.99, 3, "Software Engineering"],
]

# Write data
for row in books:
    ws.append(row)

print(f"Created workbook with {len(books)} rows")

# ============================================================================
# EXAMPLE 2: Formatting and Styling
# ============================================================================
print("\nEXAMPLE 2: Formatting and Styling")
print("-" * 70)

# Header styling
header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
header_alignment = Alignment(horizontal='center', vertical='center')

# Apply header styling
for cell in ws[1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment

print("Applied header styling")

# Data styling
data_alignment = Alignment(horizontal='left', vertical='center')
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Apply to all cells
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    for cell in row:
        cell.alignment = data_alignment
        cell.border = border
        
        # Format price column (column C)
        if cell.column == 3:
            cell.number_format = '$#,##0.00'

print("Applied data styling and number formatting")

# ============================================================================
# EXAMPLE 3: Auto-sizing Columns
# ============================================================================
print("\n📏 EXAMPLE 3: Auto-sizing Columns")
print("-" * 70)

for column_cells in ws.columns:
    length = max(len(str(cell.value)) for cell in column_cells)
    ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 2

print("Auto-sized all columns")

# ============================================================================
# EXAMPLE 4: Adding Formulas
# ============================================================================
print("\nEXAMPLE 4: Adding Formulas")
print("-" * 70)

# Add total row
ws.append([])
total_row = ws.max_row + 1
ws[f'A{total_row}'] = "TOTAL"
ws[f'A{total_row}'].font = Font(bold=True)

# Sum formula for Price column
ws[f'C{total_row}'] = f'=SUM(C2:C{total_row-2})'
ws[f'C{total_row}'].font = Font(bold=True)
ws[f'C{total_row}'].number_format = '$#,##0.00'

# Sum formula for Stock column
ws[f'D{total_row}'] = f'=SUM(D2:D{total_row-2})'
ws[f'D{total_row}'].font = Font(bold=True)

print("Added sum formulas")

# ============================================================================
# EXAMPLE 5: Multiple Sheets
# ============================================================================
print("\nEXAMPLE 5: Multiple Sheets")
print("-" * 70)

# Create summary sheet
summary = wb.create_sheet("Summary")
summary['A1'] = "Report Summary"
summary['A1'].font = Font(size=14, bold=True)

summary['A3'] = "Generated Date:"
summary['B3'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

summary['A4'] = "Total Books:"
summary['B4'] = len(books) - 1  # Exclude header

summary['A5'] = "Total Value:"
summary['B5'] = f'=Books Report!C{total_row}'
summary['B5'].number_format = '$#,##0.00'

# Style summary
for row in summary['A1:B5']:
    for cell in row:
        cell.border = border

print("Created summary sheet")

# ============================================================================
# EXAMPLE 6: Freezing Panes
# ============================================================================
print("\nEXAMPLE 6: Freezing Panes")
print("-" * 70)

# Freeze first row (header)
ws.freeze_panes = 'A2'

print("Frozen header row")

# ============================================================================
# EXAMPLE 7: Advanced Formatting
# ============================================================================
print("\nEXAMPLE 7: Advanced Formatting")
print("-" * 70)

# Highlight low stock (< 5)
yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

for row in ws.iter_rows(min_row=2, max_row=ws.max_row-2, min_col=4, max_col=4):
    for cell in row:
        if cell.value and cell.value < 5:
            cell.fill = yellow_fill
            cell.font = Font(bold=True)

print("Highlighted low stock items")

# ============================================================================
# Save File
# ============================================================================
print("\nSaving Excel File")
print("-" * 70)

filename = "books_report.xlsx"
wb.save(filename)

print(f"Saved to: {filename}")

# ============================================================================
# Summary
# ============================================================================
print("\n" + "=" * 70)
print("EXCEL EXPORT SUMMARY")
print("=" * 70)
print(f"Workbook: {len(wb.sheetnames)} sheets")
print(f"Data rows: {len(books) - 1}")
print(f"Columns: {ws.max_column}")
print(f"Features:")
print("  1. Header styling")
print("  2. Data formatting")
print("  3. Auto-sized columns")
print("  4. Formulas")
print("  5. Multiple sheets")
print("  6. Frozen panes")
print("  7. Conditional formatting")
print(f"\nFile saved: {filename}")
print("=" * 70)