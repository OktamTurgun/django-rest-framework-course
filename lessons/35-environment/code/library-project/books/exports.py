"""
Excel Export Utilities
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from datetime import datetime


class ExcelExporter:
    """Excel export utility class"""
    
    @staticmethod
    def export_books(queryset, filename="books_export.xlsx"):
        """
        Export books to Excel file
        
        Args:
            queryset: Book queryset
            filename: Output filename
            
        Returns:
            HttpResponse with Excel file
        """
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Books"
        
        # Define headers
        headers = [
            'ID', 'Title', 'Author', 'ISBN', 'Price', 
            'Stock', 'Pages', 'Language', 'Published Date', 'Genres'
        ]
        
        # Write headers
        ws.append(headers)
        
        # Style headers
        header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Write data
        for book in queryset.select_related('author').prefetch_related('genres'):
            author_name = book.author.name if book.author else 'N/A'
            genres = ', '.join([g.name for g in book.genres.all()])
            published = book.published_date.strftime('%Y-%m-%d') if book.published_date else 'N/A'
            
            ws.append([
                book.id,
                book.title,
                author_name,
                book.isbn_number,
                float(book.price),
                book.stock,
                book.pages,
                book.language,
                published,
                genres or 'N/A'
            ])
        
        # Style data cells
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical='center')
                
                # Format price column
                if cell.column == 5:  # Price column
                    cell.number_format = '$#,##0.00'
        
        # Auto-size columns
        for column_cells in ws.columns:
            length = max(len(str(cell.value or '')) for cell in column_cells)
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(length + 2, 50)
        
        # Freeze header row
        ws.freeze_panes = 'A2'
        
        # Add summary sheet
        ExcelExporter._add_summary_sheet(wb, queryset)
        
        # Create HTTP response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response
    
    @staticmethod
    def _add_summary_sheet(wb, queryset):
        """Add summary statistics sheet"""
        from django.db.models import Count, Avg, Sum, Min, Max
        
        ws = wb.create_sheet("Summary")
        
        # Title
        ws['A1'] = "Books Export Summary"
        ws['A1'].font = Font(size=14, bold=True)
        
        # Statistics
        stats = queryset.aggregate(
            total=Count('id'),
            avg_price=Avg('price'),
            min_price=Min('price'),
            max_price=Max('price'),
            total_stock=Sum('stock'),
        )
        
        ws['A3'] = "Report Date:"
        ws['B3'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        ws['A4'] = "Total Books:"
        ws['B4'] = stats['total']
        
        ws['A5'] = "Average Price:"
        ws['B5'] = float(stats['avg_price']) if stats['avg_price'] else 0
        ws['B5'].number_format = '$#,##0.00'
        
        ws['A6'] = "Min Price:"
        ws['B6'] = float(stats['min_price']) if stats['min_price'] else 0
        ws['B6'].number_format = '$#,##0.00'
        
        ws['A7'] = "Max Price:"
        ws['B7'] = float(stats['max_price']) if stats['max_price'] else 0
        ws['B7'].number_format = '$#,##0.00'
        
        ws['A8'] = "Total Stock:"
        ws['B8'] = stats['total_stock'] or 0
        
        # Style
        for row in ws['A1:B8']:
            for cell in row:
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
        
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20